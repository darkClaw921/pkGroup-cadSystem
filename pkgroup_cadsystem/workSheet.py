from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from dataclasses import dataclass
from workBitrix import create_product
import requests
from dotenv import load_dotenv
import os
load_dotenv()
PORT=os.getenv('PORT')
HOST=os.getenv('HOST')
fileName = 'list_28 (2).xlsx'
# head = 229
wbR = load_workbook(filename=fileName, read_only=True)

def send_log(message, level='INFO'):
    requests.post(f'http://{HOST}:{PORT}/logs', json={'log_entry': message, 'log_level': level})
title=('Страна', 'Менеджер', 'Наименование', 'SKU', 'Material Type', 'Product Group', 'Contract Term', 'Amount', 'Цена SRP', 'ASDK', 'ACV', 'Total', 'Цена KZT', 'Base %', 'VAR price', 'add')

@dataclass
class Product:
    country:str=0
    manager:str=1
    name:str=2
    sku:str=3
    material_type:str=4
    product_group:str=5
    contract_term:str=6
    amount:str=7
    price_srp:str=8
    asdk:str=9
    acv:str=10
    total:str=11
    price_kzt:str=12
    base_percent:str=13
    var_price:str=14
    add:str=15

    

def prepare_age(text: str) -> str:
    try:
        age = text.split(',')[1].split(' ')[1]
    except:
        age = None
    return age

def get_values(row: int):

    """Получаем масив ячеек по строке"""
    sheet_ranges = wbR['list_28 (2)']

    # for col in sheet_ranges.iter_cols(max_col=16, min_row=row, max_row=row, values_only=True):
    for row in sheet_ranges.iter_rows(max_col=16, min_row=1, max_row=20, values_only=True):
        # print(prepare_age(row[1]))
        
        print(row)
        fields={
            "NAME": row[Product.name],
            "PRICE": row[Product.price_kzt],
            "CURRENCY_ID": "KZT",
            "MEASURE": 5,
            "VAT_ID": 1,
            "VAT_INCLUDED": "Y",
            "SECTION_ID": 1,
            "ACTIVE": "Y",
            "DESCRIPTION": row[Product.add],
            "PROPERTY_1": row[Product.sku],
            "PROPERTY_2": row[Product.material_type],
            "PROPERTY_3": row[Product.product_group],
            "PROPERTY_4": row[Product.contract_term],
            "PROPERTY_5": row[Product.amount],
            "PROPERTY_6": row[Product.price_srp],
            "PROPERTY_7": row[Product.asdk],
            "PROPERTY_8": row[Product.acv],
            "PROPERTY_9": row[Product.total],
            "PROPERTY_10": row[Product.base_percent],
            "PROPERTY_11": row[Product.var_price],
            "PROPERTY_12": row[Product.add],
        }
        print(fields)
        # send_log(f'Создание товара {fields["NAME"]}')
        send_log(fields, 'DEBUG')
        # create_product(fields)
        # return row

        # for cell in row:

if __name__ == '__main__':
    get_values(1)

    